import pandas as pd
import os
import glob
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- CONFIGURACION DE ENCABEZADOS ---
DEFAULT_CONFIG = {
    # Encabezados de entrada CSV/Excel
    'COL_CARD': 'TARJETA',
    'COL_OP': 'NUM OPE',
    'COL_AMOUNT': 'IMP VISA',
    'COL_RECUPERAR': 'RECUPERAR',
    
    # Columnas Tecnicas Internas
    'AMT_FLOAT': 'Amt_Float',
    'ACCOUNTING_REF': 'Accounting_Ref',
    
    # Encabezados del Reporte de Salida
    'OUT_DEBTOR_FILE': 'ARCHIVO DEUDOR',
    'OUT_CREDIT_NOTE': 'NOTA DE CREDITO',
    'OUT_MATCHED_AMT': 'MONTO CONCILIADO',
    'OUT_STATUS': 'Estado',
    'OUT_TYPE': 'Tipo',
    'OUT_DEBTOR_FILE_COL': 'Archivo_Deudor'
}

# Keep these for backward compatibility of external scripts referencing them, 
# but internally use self.config
COL_CARD = DEFAULT_CONFIG['COL_CARD']
COL_OP = DEFAULT_CONFIG['COL_OP']
COL_AMOUNT = DEFAULT_CONFIG['COL_AMOUNT']
COL_RECUPERAR = DEFAULT_CONFIG['COL_RECUPERAR']
AMT_FLOAT = DEFAULT_CONFIG['AMT_FLOAT']
ACCOUNTING_REF = DEFAULT_CONFIG['ACCOUNTING_REF']

DEBT_PATTERN = '*m2d-recu*.xlsx'
CREDIT_PATTERN = '*m6d-dev*.xlsx'
CREDIT_PATTERN_VFF = '*m6d-dev_vff*.xlsx'
DEFAULT_FOLDER_PATH = './accounting_files'

class Conciliator:
    """
    Maneja el proceso de conciliacion entre archivos Deudores (M2D-RECU) y Acreedores (M6D-DEV).
    """
    
    def __init__(self, folder_path=DEFAULT_FOLDER_PATH, config=None):
        self.folder_path = folder_path
        self.config = DEFAULT_CONFIG.copy()
        if config:
            self.config.update(config)
            
        self.df_debt = pd.DataFrame()
        self.df_credit = pd.DataFrame()
        self.df_credit_vff = pd.DataFrame()
        self.debt_files = {}
        self.credit_files = {}
        self.credit_vff_files = {}
        self.merged = pd.DataFrame()
        self.merged_vff = pd.DataFrame()
        
        # Results
        self.pending_claims = pd.DataFrame()
        self.unexpected_refunds = pd.DataFrame()
        self.variance_report = pd.DataFrame()
        self.fully_reconciled = pd.DataFrame()
        self.net_balanced = pd.DataFrame()
        self.fully_reconciled_credits = pd.DataFrame()  # Credit-perspective reconciliation
        self.vff_debtor_notes = pd.DataFrame()  # Negative VFF differences (error)
        self.vff_acreedoras = pd.DataFrame()     # Positive VFF differences (credit notes)
        self.m6d_sin_match = pd.DataFrame()      # Orphaned M6D credits (no matching M2D)
        self.vff_abnormal = pd.DataFrame()       # Fatal crashes during VFF matching
        
        # State
        self.merged_keys = set()
        self.bad_credit_keys = set()

    # =========================================================================
    # 1. HELPER FUNCTIONS
    # =========================================================================
    def get_standardized_name(self, filepath):
        """Standardizes filename to M2D-RECU <DATE> or M6D-DEV <DATE>"""
        filename = os.path.basename(filepath)
        name_lower = filename.lower()
        
        date_match = re.search(r'(\d+[\.-]\d+[\.-]\d+)', name_lower)
        date_str = date_match.group(1) if date_match else "NO_DATE"
    # =========================================================================
    # 1.1. STANDARDIZE NAME
    # =========================================================================
        if 'm2d-recu' in name_lower:
            return f"M2D-RECU {date_str}"
        elif 'm6d-dev_vff' in name_lower:
            return f"ACREEDORA {date_str}"
        elif 'm6d-dev' in name_lower:
            return f"M6D-DEV {date_str}"
        else:
            return f"UNKNOWN {filename}"

    # =========================================================================
    # 2. DATA LOADING
    # =========================================================================
    def load_data(self):
        """Loads all data from the configured folder path."""
        self.df_debt, self.debt_files = self._load_pile(DEBT_PATTERN, "DEBT")
        self.df_credit, self.credit_files = self._load_pile(CREDIT_PATTERN, "CREDIT")
        self.df_credit_vff, self.credit_vff_files = self._load_vff_pile()
        
        has_any_credit = not self.df_credit.empty or not self.df_credit_vff.empty
        if self.df_debt.empty or not has_any_credit:
            print("Detenido: Faltan datos.")
            return False
        return True

    def _load_pile(self, pattern, label):
        """Internal helper to load files matching a pattern (parallel I/O)."""
        files = glob.glob(os.path.join(self.folder_path, pattern))
        filter_keyword = 'm2d-recu' if label == "DEBT" else 'm6d-dev'
        files = [f for f in files if filter_keyword in os.path.basename(f).lower()]
        # Exclude VFF files from regular credit pile
        if label == "CREDIT":
            files = [f for f in files if 'm6d-dev_vff' not in os.path.basename(f).lower()]
        
        all_dfs = []
        individual_files = {}
        print(f"Cargando {len(files)} archivos para {label}...")

        # Parallel file loading (with ordered array collection)
        results = [None] * len(files)
        with ThreadPoolExecutor(max_workers=max(1, min(len(files), os.cpu_count() or 4))) as executor:
            future_to_index = {executor.submit(self._process_single_file, f): i for i, f in enumerate(files)}
            for future in as_completed(future_to_index):
                i = future_to_index[future]
                try:
                    results[i] = future.result()
                except Exception as e:
                    print(f"  [ERROR] {os.path.basename(files[i])}: {e}")

        for df in results:
            if df is not None:
                std_name = df[ACCOUNTING_REF].iloc[0]
                all_dfs.append(df)
                individual_files[std_name] = df

        combined = pd.concat(all_dfs, ignore_index=True) if all_dfs else pd.DataFrame()
        return combined, individual_files

    def _load_vff_pile(self):
        """
        Loads VFF (M6D-DEV_VFF) files and computes credit notes from paired operations.
        
        VFF Business Rules:
        - Each VFF file contains paired operations (same TARJETA + NUM OPE)
        - Within each pair: row 1 = debtor import, row 2 = creditor import
        - Credit note = creditor_amount - debtor_amount
        - If difference is negative → debtor note (stored as error in self.vff_debtor_notes)
        - Unpaired operations in one file are matched across other VFF files
        """
        files = glob.glob(os.path.join(self.folder_path, CREDIT_PATTERN_VFF))
        files = [f for f in files if 'm6d-dev_vff' in os.path.basename(f).lower()]
        
        all_dfs = []
        individual_files = {}
        print(f"Cargando {len(files)} archivos para CREDIT_VFF...")
        
        # Parallel file loading (with ordered array collection)
        results = [None] * len(files)
        with ThreadPoolExecutor(max_workers=max(1, min(len(files), os.cpu_count() or 4))) as executor:
            future_to_index = {executor.submit(self._process_single_file, f): i for i, f in enumerate(files)}
            for future in as_completed(future_to_index):
                i = future_to_index[future]
                try:
                    results[i] = future.result()
                except Exception as e:
                    print(f"  [ERROR] {os.path.basename(files[i])}: {e}")
                    
        for df in results:
            if df is not None:
                std_name = df[ACCOUNTING_REF].iloc[0]
                all_dfs.append(df)
                individual_files[std_name] = df
        
        if not all_dfs:
            print("  No se encontraron archivos VFF.")
            return pd.DataFrame(), {}
        
        # Concatenate all VFF data for cross-file matching
        combined = pd.concat(all_dfs, ignore_index=True)
        computed_credits = self._compute_vff_pairs(combined)
        
        return computed_credits, individual_files

    def _compute_vff_pairs(self, df_vff):
        """
        Groups VFF rows by (TARJETA, NUM OPE) and computes credit notes.
        Row 1 of each pair = debtor, Row 2 = creditor.
        Returns a DataFrame of computed credit notes (one row per pair).
        """
        if df_vff.empty:
            return pd.DataFrame()
        
        credit_rows = []
        debtor_note_rows = []
        
        grouped = df_vff.groupby([COL_CARD, COL_OP])
        
        abnormal_rows = []
        
        for (card, op), group in grouped:
            try:
                group = group.reset_index(drop=True)
                
                if len(group) == 2:
                    # Normal pair: row 0 = debtor, row 1 = creditor
                    debtor_amt = group[AMT_FLOAT].iloc[0]
                    creditor_amt = group[AMT_FLOAT].iloc[1]
                    difference = creditor_amt - debtor_amt
                    
                    row = {
                        COL_CARD: card,
                        COL_OP: op,
                        AMT_FLOAT: abs(difference),
                        ACCOUNTING_REF: group[ACCOUNTING_REF].iloc[0],
                        'VFF_Debtor_Amt': debtor_amt,
                        'VFF_Creditor_Amt': creditor_amt,
                        'VFF_Difference': difference,
                        'VFF_Source_Files': ', '.join(group[ACCOUNTING_REF].unique()),
                    }
                    # Copy RECUPERAR if available
                    if COL_RECUPERAR in group.columns:
                        row[COL_RECUPERAR] = group[COL_RECUPERAR].iloc[0]
                    
                    if difference < 0:
                        row['VFF_Note_Type'] = 'DEBTOR_NOTE (NEGATIVE)'
                        debtor_note_rows.append(row)
                        print(f"  ⚠ NOTA DEUDORA VFF: {card}/{op} dif={difference:.2f}")
                    else:
                        row['VFF_Note_Type'] = 'CREDIT_NOTE'
                        credit_rows.append(row)
                        
                elif len(group) == 1:
                    # Single operation with no pair — should not normally happen
                    print(f"  ⚠ VFF SIN PAR: {card}/{op} en {group[ACCOUNTING_REF].iloc[0]} (cruce entre archivos)")
                    # This was already handled by cross-file concatenation;
                    # if still single after concat, it's truly unpaired
                    row = {
                        COL_CARD: card,
                        COL_OP: op,
                        AMT_FLOAT: group[AMT_FLOAT].iloc[0],
                        ACCOUNTING_REF: group[ACCOUNTING_REF].iloc[0],
                        'VFF_Debtor_Amt': group[AMT_FLOAT].iloc[0],
                        'VFF_Creditor_Amt': 0.0,
                        'VFF_Difference': 0.0,
                        'VFF_Source_Files': group[ACCOUNTING_REF].iloc[0],
                        'VFF_Note_Type': 'UNPAIRED',
                    }
                    if COL_RECUPERAR in group.columns:
                        row[COL_RECUPERAR] = group[COL_RECUPERAR].iloc[0]
                    debtor_note_rows.append(row)
                else:
                    # More than 2 rows for same key — unexpected
                    print(f"  ⚠ VFF ANOMALIA: {card}/{op} tiene {len(group)} filas")
                    # Take the first two as the pair
                    debtor_amt = group[AMT_FLOAT].iloc[0]
                    creditor_amt = group[AMT_FLOAT].iloc[1]
                    difference = creditor_amt - debtor_amt
                    row = {
                        COL_CARD: card,
                        COL_OP: op,
                        AMT_FLOAT: abs(difference),
                        ACCOUNTING_REF: group[ACCOUNTING_REF].iloc[0],
                        'VFF_Debtor_Amt': debtor_amt,
                        'VFF_Creditor_Amt': creditor_amt,
                        'VFF_Difference': difference,
                        'VFF_Source_Files': ', '.join(group[ACCOUNTING_REF].unique()),
                        'VFF_Note_Type': 'ANOMALY_MULTI_ROW',
                    }
                    if COL_RECUPERAR in group.columns:
                        row[COL_RECUPERAR] = group[COL_RECUPERAR].iloc[0]
                    debtor_note_rows.append(row)
            except Exception as e:
                # Catch any severe crash gracefully and log to Abnormal sheet
                print(f"  [ERROR GRAVE] Fallo inesperado en VFF {card}/{op}: {e}")
                for _, r in group.iterrows():
                    abnormal_rows.append({
                        'Archivo_Origen': r.get(ACCOUNTING_REF, 'Desconocido'),
                        'Tarjeta': card,
                        'Operacion': op,
                        'Monto_Original': r.get(AMT_FLOAT, 0.0),
                        'Motivo_Crash': str(e)
                    })
        
        self.vff_abnormal = pd.DataFrame(abnormal_rows) if abnormal_rows else pd.DataFrame()
        if not self.vff_abnormal.empty:
            print(f"  ⚠ {len(self.vff_abnormal)} operaciones causaron fallos criticos (movidos a VFF_Error_Fatal)")
        
        # Store debtor notes (negative differences + unpaired + anomalies)
        self.vff_debtor_notes = pd.DataFrame(debtor_note_rows) if debtor_note_rows else pd.DataFrame()
        if not self.vff_debtor_notes.empty:
            print(f"  ⚠ Se encontraron {len(self.vff_debtor_notes)} notas deudoras VFF / errores")
        
        # Store valid credit notes
        self.vff_acreedoras = pd.DataFrame(credit_rows) if credit_rows else pd.DataFrame()
        if not self.vff_acreedoras.empty:
            print(f"  ✓ Se calcularon {len(self.vff_acreedoras)} notas de credito VFF (acreedoras)")
        
        return self.vff_acreedoras

    def _process_single_file(self, filepath):
        """Reads and cleans a single Excel file."""
        # Load as String
        df = pd.read_excel(filepath, dtype=str)
        
        if COL_CARD in df.columns and COL_OP in df.columns:
            df[COL_CARD] = df[COL_CARD].replace(r'^\s*$', pd.NA, regex=True)
            df[COL_OP] = df[COL_OP].replace(r'^\s*$', pd.NA, regex=True)
            
            # Drop empty trailing rows
            df = df.dropna(subset=[COL_CARD, COL_OP], how='all')
        
        # Standardized Name
        std_name = self.get_standardized_name(filepath)
        df[ACCOUNTING_REF] = std_name
        
        # Clean Keys
        if COL_CARD in df.columns and COL_OP in df.columns:
            df[COL_CARD] = df[COL_CARD].str.strip()
            df[COL_OP] = df[COL_OP].str.strip()
        else:
            print(f"  [OMITIDO] {std_name} falta encabezado de Tarjeta u Operacion.")
            return None
        
        # Clean Amount
        if COL_AMOUNT in df.columns:
            clean_amt = df[COL_AMOUNT].astype(str).str.replace(r'[^\d.-]', '', regex=True)
            df[AMT_FLOAT] = pd.to_numeric(clean_amt, errors='coerce').fillna(0.0)
        
        # Clean RECUPERAR
        if COL_RECUPERAR in df.columns:
            df[COL_RECUPERAR] = df[COL_RECUPERAR].astype(str).str.strip().str.upper()
        else:
            df[COL_RECUPERAR] = 'SI'

        return df

    # =========================================================================
    # 3. VALIDATION
    # =========================================================================
    def validate(self):
        """Orchestrates validation checks."""
        print("Verificando archivos duplicados dentro de cada categoria...")
        intra_issues = self._check_intra_pile_duplicates(self.debt_files, "DEBT") + \
                       self._check_intra_pile_duplicates(self.credit_files, "CREDIT")
        
        if intra_issues:
            self._print_validation_errors("DETECCION DE DUPLICADOS INTRA-CATEGORIA", intra_issues)
            return False

        print("Validando que los archivos no sean duplicados...")
        dups = self._validate_files_are_different(self.df_debt, self.df_credit)
        if dups:
            self._print_validation_errors("DETECCION DE ARCHIVOS DUPLICADOS", dups)
            return False

        print("Ejecutando verificaciones de calidad de datos...")
        if not self._run_quality_checks():
            return False
            
        return True

    def _print_validation_errors(self, title, errors):
        print(f"\n{'='*60}\n⚠️  {title} ⚠️\n{'='*60}")
        for e in errors: print(f"  ❌ {e}")
        print("\nConciliacion ABORTADA.\n")

    def _check_intra_pile_duplicates(self, individual_files, label):
        issues = []
        # Optimization: Only compare files with same number of rows
        from collections import defaultdict
        files_by_len = defaultdict(list)
        
        for name, df in individual_files.items():
            files_by_len[len(df)].append(name)
            
        for length, names in files_by_len.items():
            if len(names) < 2: continue
            
            # Compare only within this group
            for i in range(len(names)):
                for j in range(i + 1, len(names)):
                    name1, name2 = names[i], names[j]
                    df1, df2 = individual_files[name1], individual_files[name2]
                    
                    keys1 = set(zip(df1[self.config['COL_CARD']], df1[self.config['COL_OP']]))
                    keys2 = set(zip(df2[self.config['COL_CARD']], df2[self.config['COL_OP']]))
                    
                    if keys1 == keys2:
                        compare_cols = [col for col in df1.columns if col not in [self.config['ACCOUNTING_REF']]]
                        df1_sorted = df1[compare_cols].sort_values(by=[self.config['COL_CARD'], self.config['COL_OP']]).reset_index(drop=True)
                        df2_sorted = df2[compare_cols].sort_values(by=[self.config['COL_CARD'], self.config['COL_OP']]).reset_index(drop=True)
                        
                        if df1_sorted.equals(df2_sorted):
                            issues.append(f"ARCHIVOS {label} DUPLICADOS: '{name1}' y '{name2}' contienen datos IDENTICOS!")
                        else:
                            issues.append(f"ARCHIVOS {label} SOSPECHOSOS: '{name1}' y '{name2}' tienen operaciones identicas pero montos diferentes!")
                    else:
                        overlap = keys1 & keys2
                        overlap_pct = len(overlap) / max(len(keys1), 1) * 100
                        if overlap_pct > 90:
                            issues.append(f"ADVERTENCIA {label}: '{name1}' y '{name2}' comparten {overlap_pct:.1f}% de operaciones!")
        return issues

    def _validate_files_are_different(self, df1, df2):
        issues = []
        compare_cols = [col for col in df1.columns if col not in [ACCOUNTING_REF, AMT_FLOAT]]
        
        # 1. Exact DataFrame Match
        if set(compare_cols) == set([col for col in df2.columns if col not in [ACCOUNTING_REF, AMT_FLOAT]]):
            df1_cmp = df1[compare_cols].reset_index(drop=True)
            df2_cmp = df2[compare_cols].reset_index(drop=True)
            if len(df1_cmp) == len(df2_cmp):
                df1_s = df1_cmp.sort_values(by=compare_cols).reset_index(drop=True)
                df2_s = df2_cmp.sort_values(by=compare_cols).reset_index(drop=True)
                if df1_s.equals(df2_s):
                    issues.append("COINCIDENCIA EXACTA: Los archivos DEUDOR y ACREEDOR contienen datos identicos!")

        # 2. Key Overlap
        debt_keys = set(zip(df1[COL_CARD], df1[COL_OP]))
        credit_keys = set(zip(df2[COL_CARD], df2[COL_OP]))
        overlap_pct = len(debt_keys & credit_keys) / max(len(debt_keys), 1) * 100
        if overlap_pct > 95 and len(debt_keys) == len(credit_keys):
            issues.append(f"SOSPECHOSO: {overlap_pct:.1f}% de coincidencia de claves con mismo numero de filas!")

        # 3. Amount Fingerprint
        if AMT_FLOAT in df1.columns and AMT_FLOAT in df2.columns:
            if (abs(df1[AMT_FLOAT].sum() - df2[AMT_FLOAT].sum()) < 0.01 and 
                abs(df1[AMT_FLOAT].mean() - df2[AMT_FLOAT].mean()) < 0.01 and
                len(df1) == len(df2)):
                issues.append("SOSPECHOSO: Suma, promedio y cantidad de filas identicos!")

        # 4. Source Type Check
        debt_sources = {s.split()[0] for s in df1[ACCOUNTING_REF].unique()}
        credit_sources = {s.split()[0] for s in df2[ACCOUNTING_REF].unique()}
        if debt_sources == credit_sources:
            issues.append(f"ADVERTENCIA: Ambas fuentes son tipo '{debt_sources}' - se esperaban tipos diferentes!")
            
        return issues

    def _run_quality_checks(self):
        all_warnings, all_errors = [], []
        
        for f, df in self.debt_files.items():
            w, e = self._check_data_quality(df, f"DEBT ({f})")
            all_warnings.extend(w); all_errors.extend(e)
        for f, df in self.credit_files.items():
            w, e = self._check_data_quality(df, f"CREDIT ({f})")
            all_warnings.extend(w); all_errors.extend(e)

        if all_warnings:
            print(f"\n{'-'*60}\n⚠️  ADVERTENCIAS DE CALIDAD DE DATOS\n{'-'*60}")
            for w in all_warnings: print(f"  ⚠ {w}")
        
        if all_errors:
            self._print_validation_errors("ERRORES DE CALIDAD DE DATOS", all_errors)
            return False
        return True

    def _check_data_quality(self, df, label):
        warnings = []
        errors = []
        
        if AMT_FLOAT in df.columns:
            if (df[AMT_FLOAT] < 0).any(): warnings.append(f"{label}: Se encontraron montos negativos")
            if (df[AMT_FLOAT] == 0).any(): warnings.append(f"{label}: Se encontraron transacciones con monto cero")
            
            # Outliers (>3 std)
            if len(df) > 10:
                mean, std = df[AMT_FLOAT].mean(), df[AMT_FLOAT].std()
                if std > 0 and not df[df[AMT_FLOAT] > mean + 3*std].empty:
                    warnings.append(f"{label}: Se encontraron montos inusualmente grandes")

        if COL_CARD in df.columns:
            empty = df[COL_CARD].isna().sum() + (df[COL_CARD] == '').sum()
            if empty > 0: errors.append(f"{label}: {empty} filas con numeros de Tarjeta vacios")
            
        if COL_OP in df.columns:
            empty = df[COL_OP].isna().sum() + (df[COL_OP] == '').sum()
            if empty > 0: errors.append(f"{label}: {empty} filas con numeros de Operacion vacios")
            
        # Duplicates within file
        if COL_CARD in df.columns and COL_OP in df.columns:
            dups = df.groupby([COL_CARD, COL_OP, ACCOUNTING_REF]).size()
            if (dups > 1).any():
                warnings.append(f"{label}: Se encontraron combinaciones de clave duplicadas dentro del mismo archivo")

        return warnings, errors

    # =========================================================================
    # 4. MATCHING & ANALYSIS
    # =========================================================================
    def match_transactions(self):
        print("Conciliando Transacciones...")
        if not self.df_credit.empty:
            self.merged = pd.merge(
                self.df_debt, 
                self.df_credit, 
                on=[COL_CARD, COL_OP], 
                how='inner', 
                suffixes=('_DEBT', '_CREDIT')
            )
        else:
            self.merged = pd.DataFrame()
        
        # Match VFF credits against M2D debts
        self._match_vff_transactions()
        
        if self.merged.empty and self.merged_vff.empty:
            print("No se encontraron coincidencias.")
            return False
            
        success = self._check_orphans()
        
        # Integrate elegantly so they appear naturally in unified summaries and analytical trackers
        if not self.merged_vff.empty:
            if self.merged.empty:
                self.merged = self.merged_vff.copy()
            else:
                self.merged = pd.concat([self.merged, self.merged_vff], ignore_index=True)
                
        return success

    def _match_vff_transactions(self):
        """Matches VFF computed credit notes against M2D debt files."""
        if self.df_credit_vff.empty:
            print("  No hay creditos VFF para conciliar.")
            return
        
        print("Conciliando creditos VFF contra deudas M2D...")
        self.merged_vff = pd.merge(
            self.df_debt,
            self.df_credit_vff,
            on=[COL_CARD, COL_OP],
            how='inner',
            suffixes=('_DEBT', '_CREDIT')
        )
        
        if not self.merged_vff.empty:
            print(f"  ✓ Se conciliaron {len(self.merged_vff)} creditos VFF con deudas M2D")
        
        # Unmatched VFF credits → unexpected refunds
        if not self.df_credit_vff.empty:
            vff_keys = set(zip(self.df_credit_vff[COL_CARD], self.df_credit_vff[COL_OP]))
            matched_vff_keys = set(zip(self.merged_vff[COL_CARD], self.merged_vff[COL_OP])) if not self.merged_vff.empty else set()
            unmatched_vff_keys = vff_keys - matched_vff_keys
            
            if unmatched_vff_keys:
                print(f"  ⚠ {len(unmatched_vff_keys)} creditos VFF sin coincidencia M2D → Devoluciones Inesperadas")
                self.df_credit_vff['temp_key'] = list(zip(self.df_credit_vff[COL_CARD], self.df_credit_vff[COL_OP]))
                unmatched_vff = self.df_credit_vff[self.df_credit_vff['temp_key'].isin(unmatched_vff_keys)].copy()
                unmatched_vff.drop(columns=['temp_key'], inplace=True)
                # Add VFF source marker
                unmatched_vff['VFF_Source'] = 'VFF_UNMATCHED'
                self.unexpected_refunds = pd.concat([self.unexpected_refunds, unmatched_vff], ignore_index=True)
                self.df_credit_vff.drop(columns=['temp_key'], inplace=True)

    def _check_orphans(self):
        print("Analizando registros sin conciliar...")
        merged_keys = set(zip(self.merged[COL_CARD], self.merged[COL_OP])) if not self.merged.empty else set()
        
        # Also include VFF matched keys
        vff_matched_keys = set(zip(self.merged_vff[COL_CARD], self.merged_vff[COL_OP])) if not self.merged_vff.empty else set()
        all_matched_keys = merged_keys | vff_matched_keys
        
        # Check orphans for regular credits (NON-BLOCKING — alert + M6D SIN MATCH sheet)
        if not self.df_credit.empty:
            credit_keys = set(zip(self.df_credit[COL_CARD], self.df_credit[COL_OP]))
            orphaned_credit_keys = credit_keys - merged_keys
            
            if orphaned_credit_keys:
                print(f"\n{'='*60}")
                print(f"⚠️  M6D SIN MATCH: {len(orphaned_credit_keys)} creditos SIN deuda correspondiente")
                print(f"{'='*60}")
                
                # Use fast vectorised merge for extracting orphans
                keys_df = pd.DataFrame(list(orphaned_credit_keys), columns=[COL_CARD, COL_OP])
                orphans_df = self.df_credit.merge(keys_df, on=[COL_CARD, COL_OP], how='inner').copy()
                problem_files = orphans_df[ACCOUNTING_REF].unique()
                
                print("\n  📂 ARCHIVOS M6D CON OPERACIONES SIN CONCILIAR:")
                for f in problem_files:
                    file_orphan_count = len(orphans_df[orphans_df[ACCOUNTING_REF] == f])
                    print(f"     - {f}: {file_orphan_count} operacion(es) sin conciliar")
                
                # Store for export — include origin file name
                orphans_df.rename(columns={ACCOUNTING_REF: 'Archivo_Origen'}, inplace=True)
                self.m6d_sin_match = orphans_df
                
                print(f"\n  ℹ️  Se exportaran a la hoja 'M6D SIN MATCH'.")
                print(f"  La conciliacion continuara.\n")
        
        # VFF orphaned credits are NOT blocking — already funneled to unexpected refunds
        
        debt_keys = set(zip(self.df_debt[COL_CARD], self.df_debt[COL_OP]))
        orphaned_debt_keys = debt_keys - all_matched_keys
        if orphaned_debt_keys:
            print(f"\n📊 DEUDAS SIN CONCILIAR: {len(orphaned_debt_keys)} (Informativo)")
        else:
            print("✓ Todos los creditos conciliados con deudas (100% conciliacion).")
            
        return True

    def run_analysis(self):
        self._identify_recuperar_scenarios()
        self._analyze_variance()

    def _analyze_variance(self):
        print("Verificando varianzas de montos...")
        variance_check = self.merged.groupby(
            [f"{self.config['ACCOUNTING_REF']}_CREDIT", self.config['COL_CARD'], self.config['COL_OP'], f"{self.config['AMT_FLOAT']}_CREDIT"]
        ).agg(
            Total_Debts_Covered=(f"{self.config['AMT_FLOAT']}_DEBT", 'sum')
        ).reset_index()
        
        variance_check['Varianza'] = variance_check[f"{self.config['AMT_FLOAT']}_CREDIT"] - variance_check['Total_Debts_Covered']
        self.variance_report = variance_check[variance_check['Varianza'].abs() > 0.01].copy()
        
        if not self.variance_report.empty:
            print(f"  ⚠ Se encontraron {len(self.variance_report)} COINCIDENCIAS CON VARIANZA")
            self.variance_report['Estado'] = self.variance_report['Varianza'].apply(
                lambda x: "SOBREPAGO (Devolucion > Deudas)" if x > 0 else "PAGO INSUFICIENTE (Devolucion < Deudas)"
            )
            # Collect bad keys to exclude from strict reconciliation
            for _, row in self.variance_report.iterrows():
                self.bad_credit_keys.add((row[f"{self.config['ACCOUNTING_REF']}_CREDIT"], row[self.config['COL_CARD']], row[self.config['COL_OP']]))
                
            self.variance_report.rename(columns={
                f"{self.config['ACCOUNTING_REF']}_CREDIT": 'Archivo_Credito',
                f"{self.config['AMT_FLOAT']}_CREDIT": 'Monto_Devolucion'
            }, inplace=True)

    def _identify_recuperar_scenarios(self):
        print("Aplicando logica de negocio 'RECUPERAR'...")
        
        # 1. Pending Claims (RECUPERAR='NO' and not matched)
        self.merged_keys = set(zip(self.merged[self.config['COL_CARD']], self.merged[self.config['COL_OP']])) if not self.merged.empty else set()
        
        # Fast Anti-Join using merge instead of tuple mapping
        debt_no = self.df_debt[self.df_debt[self.config['COL_RECUPERAR']] == 'NO'].copy()
        if not debt_no.empty and self.merged_keys:
            merged_keys_df = pd.DataFrame(list(self.merged_keys), columns=[self.config['COL_CARD'], self.config['COL_OP']])
            merged_keys_df['_is_matched'] = True
            joined = debt_no.merge(merged_keys_df, on=[self.config['COL_CARD'], self.config['COL_OP']], how='left')
            self.pending_claims = joined[joined['_is_matched'].isna()].drop(columns=['_is_matched']).copy()
        else:
            self.pending_claims = debt_no if not self.merged_keys else pd.DataFrame()
        
        if not self.pending_claims.empty:
            print(f"  ⚠ Se encontraron {len(self.pending_claims)} DEUDORAS PENDIENTES")
            
        # 2. Unexpected Refunds (RECUPERAR!='NO' but matched)
        self.unexpected_refunds = self.merged[self.merged[f"{self.config['COL_RECUPERAR']}_DEBT"] != 'NO'].copy()
        
        if not self.unexpected_refunds.empty:
            print(f"  ℹ Se encontraron {len(self.unexpected_refunds)} DEVOLUCIONES INESPERADAS")

    # =========================================================================
    # 5. REPORT GENERATION
    # =========================================================================
    def generate_reports(self):
        self._generate_fully_reconciled_summary()
        self._generate_credit_reconciled_summary()
        self._generate_net_balanced_summary()

    def _generate_fully_reconciled_summary(self):
        print("Generando Resumen de Conciliacion Completa (Modo Estricto)...")
        summary_rows = []
        debt_groups = self.df_debt.groupby(self.config['ACCOUNTING_REF'])
        
        # Optimization: Pre-compute exclusions sets for O(1) lookup
        pending_exclusion_set = set(self.pending_claims[self.config['ACCOUNTING_REF']].unique()) if not self.pending_claims.empty else set()
        unexpected_exclusion_set = set(self.unexpected_refunds[f"{self.config['ACCOUNTING_REF']}_DEBT"].unique()) if not self.unexpected_refunds.empty else set()
        
        # Pre-group the merged dataframe by DEBT file where RECUPERAR is 'NO'
        # This reduces filtering from O(N^2) down to O(N) lookup !
        if not self.merged.empty:
            filtered_merged = self.merged[self.merged[f"{self.config['COL_RECUPERAR']}_DEBT"] == 'NO']
            merged_by_debt = {k: v for k, v in filtered_merged.groupby(f"{self.config['ACCOUNTING_REF']}_DEBT")}
        else:
            merged_by_debt = {}

        for filename, group in debt_groups:
            # Check Exclusions first (fastest check)
            if filename in pending_exclusion_set: continue
            if filename in unexpected_exclusion_set: continue
            
            total_no = group[group[self.config['COL_RECUPERAR']] == 'NO']
            if total_no.empty: continue
            
            # Lookup pre-filtered frame
            relevant_merged = merged_by_debt.get(filename, pd.DataFrame())
            
            # Verify 100% Match (if lengths match, everything matched since there are no duplicates and they joined strictly)
            if len(total_no) != len(relevant_merged): continue
            
            # Set intersection for variance check (faster than loop)
            current_keys = set(zip(relevant_merged[f"{self.config['ACCOUNTING_REF']}_CREDIT"], relevant_merged[self.config['COL_CARD']], relevant_merged[self.config['COL_OP']]))
            if not current_keys.isdisjoint(self.bad_credit_keys):
                continue
            
            # --- Build ticket-ready breakdown ---
            unique_ops = relevant_merged.drop_duplicates(subset=[self.config['COL_CARD'], self.config['COL_OP']])
            debtor_total = unique_ops[f"{self.config['AMT_FLOAT']}_DEBT"].sum()
            debtor_op_count = len(unique_ops)
            
            # Credit note breakdown: which credit files cover this debtor note
            creditor_breakdown = unique_ops.groupby(f"{self.config['ACCOUNTING_REF']}_CREDIT").agg(
                Credit_Amount=(f"{self.config['AMT_FLOAT']}_CREDIT", 'sum'),
                Operations=(self.config['COL_OP'], 'count')
            ).reset_index()
            
            # Add inline header for this block
            summary_rows.append({
                'Nota Deudora': 'Nota Deudora',
                'Total Deudora': 'Total Deudora',
                'Notas Acreedoras': 'Notas Acreedoras',
                'Monto Acreedor': 'Monto Acreedor'
            })
            
            first_row = True
            for _, crow in creditor_breakdown.iterrows():
                summary_rows.append({
                    'Nota Deudora': filename if first_row else '',
                    'Total Deudora': debtor_total if first_row else '',
                    'Notas Acreedoras': crow[f"{self.config['ACCOUNTING_REF']}_CREDIT"],
                    'Monto Acreedor': crow['Credit_Amount']
                })
                first_row = False
            
            # Subtotal row for this debtor file
            summary_rows.append({
                'Nota Deudora': f"SUBTOTAL {filename}",
                'Total Deudora': '',
                'Notas Acreedoras': '',
                'Monto Acreedor': creditor_breakdown['Credit_Amount'].sum()
            })
            
            # Blank line
            summary_rows.append({
                'Nota Deudora': '',
                'Total Deudora': '',
                'Notas Acreedoras': '',
                'Monto Acreedor': ''
            })
                
        self.fully_reconciled = pd.DataFrame(summary_rows)

    def _generate_credit_reconciled_summary(self):
        """Generates credit-perspective summary: each credit note → which debtor notes it covers."""
        print("Generando Resumen de Conciliacion por Nota de Credito...")
        if self.merged.empty:
            return
        
        summary_rows = []
        credit_groups = self.merged.groupby(f"{self.config['ACCOUNTING_REF']}_CREDIT")
        
        for credit_name, credit_group in credit_groups:
            unique_ops = credit_group.drop_duplicates(subset=[self.config['COL_CARD'], self.config['COL_OP']])
            credit_total = unique_ops[f"{self.config['AMT_FLOAT']}_CREDIT"].sum() if f"{self.config['AMT_FLOAT']}_CREDIT" in unique_ops.columns else 0.0
            credit_op_count = len(unique_ops)
            
            # Debtor breakdown: which debtor files this credit covers
            debtor_breakdown = unique_ops.groupby(f"{self.config['ACCOUNTING_REF']}_DEBT").agg(
                Debtor_Amount=(f"{self.config['AMT_FLOAT']}_DEBT", 'sum'),
                Operations=(self.config['COL_OP'], 'count')
            ).reset_index()
            
            # Add inline header for this block
            summary_rows.append({
                'Nota Acreedora': 'Nota Acreedora',
                'Total Acreedora': 'Total Acreedora',
                'Notas Deudoras': 'Notas Deudoras',
                'Monto Deudor': 'Monto Deudor'
            })
            
            first_row = True
            for _, drow in debtor_breakdown.iterrows():
                summary_rows.append({
                    'Nota Acreedora': credit_name if first_row else '',
                    'Total Acreedora': credit_total if first_row else '',
                    'Notas Deudoras': drow[f"{self.config['ACCOUNTING_REF']}_DEBT"],
                    'Monto Deudor': drow['Debtor_Amount']
                })
                first_row = False
            
            # Subtotal row for this credit file
            summary_rows.append({
                'Nota Acreedora': f"SUBTOTAL {credit_name}",
                'Total Acreedora': '',
                'Notas Deudoras': '',
                'Monto Deudor': debtor_breakdown['Debtor_Amount'].sum()
            })
            
            # Blank line
            summary_rows.append({
                'Nota Acreedora': '',
                'Total Acreedora': '',
                'Notas Deudoras': '',
                'Monto Deudor': ''
            })
        
        self.fully_reconciled_credits = pd.DataFrame(summary_rows)

    def _generate_net_balanced_summary(self):
        print("Verificando archivos con balance neto...")
        rows = []
        
        candidates = set(self.pending_claims[self.config['ACCOUNTING_REF']].unique()) | set(self.unexpected_refunds[f"{self.config['ACCOUNTING_REF']}_DEBT"].unique())
        
        if not self.fully_reconciled.empty:
            excluded = set(self.fully_reconciled['Nota Deudora'].dropna().unique())
            candidates = candidates - excluded
        
        # Optimization: Pre-group to avoid O(N) filtering inside loop
        pending_map = {k: v for k, v in self.pending_claims.groupby(self.config['ACCOUNTING_REF'])} if not self.pending_claims.empty else {}
        unexpected_map = {k: v for k, v in self.unexpected_refunds.groupby(f"{self.config['ACCOUNTING_REF']}_DEBT")} if not self.unexpected_refunds.empty else {}
        
        for filename in candidates:
            if filename == 'TOTAL': continue
            
            file_pending = pending_map.get(filename, pd.DataFrame())
            file_unexpected = unexpected_map.get(filename, pd.DataFrame())
            
            sum_p = file_pending[self.config['COL_AMOUNT']].astype(float).sum() if not file_pending.empty else 0.0
            sum_u = file_unexpected[f"{self.config['AMT_FLOAT']}_CREDIT"].sum() if not file_unexpected.empty else 0.0
            
            if abs(sum_p - sum_u) < 0.01 and (sum_p > 0 or sum_u > 0):
                # IT IS BALANCED
                if not file_pending.empty:
                    for _, r in file_pending.iterrows():
                        rows.append({
                            self.config['OUT_DEBTOR_FILE_COL']: filename, 
                            self.config['OUT_STATUS']: 'NET BALANCED', 
                            self.config['OUT_TYPE']: 'PENDING_CLAIM', 
                            self.config['OUT_MATCHED_AMT']: float(r[self.config['COL_AMOUNT']])
                        })
                if not file_unexpected.empty:
                    for _, r in file_unexpected.iterrows():
                        rows.append({
                            self.config['OUT_DEBTOR_FILE_COL']: filename, 
                            self.config['OUT_STATUS']: 'NET BALANCED', 
                            self.config['OUT_TYPE']: 'UNEXPECTED_REFUND', 
                            self.config['OUT_MATCHED_AMT']: r[f"{self.config['AMT_FLOAT']}_CREDIT"]
                        })
                
                # Context lookup - could be optimized further but acceptable for now
                file_matched = self.merged[
                     (self.merged[f"{self.config['ACCOUNTING_REF']}_DEBT"] == filename) & 
                     (self.merged[f"{self.config['COL_RECUPERAR']}_DEBT"] == 'NO')
                ]
                for _, r in file_matched.iterrows():
                    rows.append({
                        self.config['OUT_DEBTOR_FILE_COL']: filename, 
                        self.config['OUT_STATUS']: 'NET BALANCED', 
                        self.config['OUT_TYPE']: 'CORRECTLY_MATCHED', 
                        self.config['OUT_MATCHED_AMT']: r[f"{self.config['AMT_FLOAT']}_DEBT"]
                    })
                    
        self.net_balanced = pd.DataFrame(rows)

    def export_results(self, output_file):
        try:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # Breakdowns
                if not self.merged.empty:
                    self.merged.groupby([f"{self.config['ACCOUNTING_REF']}_DEBT", f"{self.config['ACCOUNTING_REF']}_CREDIT"]).agg(
                        Cantidad=(self.config['COL_OP'], 'count'), Monto=(f"{self.config['AMT_FLOAT']}_DEBT", 'sum')
                    ).reset_index().to_excel(writer, sheet_name='Por_Archivo_Deudor', index=False)
                    
                    self.merged.groupby([f"{self.config['ACCOUNTING_REF']}_CREDIT", f"{self.config['ACCOUNTING_REF']}_DEBT"]).agg(
                        Cantidad=(self.config['COL_OP'], 'count'), Monto=(f"{self.config['AMT_FLOAT']}_DEBT", 'sum')
                    ).reset_index().to_excel(writer, sheet_name='Por_Archivo_Credito', index=False)
                
                if not self.pending_claims.empty: self.pending_claims.to_excel(writer, sheet_name='DEUDORAS_Pendientes', index=False)
                if not self.unexpected_refunds.empty: self.unexpected_refunds.to_excel(writer, sheet_name='Devoluciones_Inesperadas', index=False)
                if not self.fully_reconciled.empty: self.fully_reconciled.to_excel(writer, sheet_name='Notas_Conciliadas', index=False, header=False)
                if not self.fully_reconciled_credits.empty: self.fully_reconciled_credits.to_excel(writer, sheet_name='Conciliado_Por_Credito', index=False, header=False)
                if not self.net_balanced.empty: self.net_balanced.to_excel(writer, sheet_name='Balance_Neto', index=False)
                if not self.variance_report.empty: self.variance_report.to_excel(writer, sheet_name='Varianzas_Monto', index=False)
                
                # M6D SIN MATCH - orphaned M6D credits with origin filenames
                if not self.m6d_sin_match.empty:
                    self.m6d_sin_match.to_excel(writer, sheet_name='M6D SIN MATCH', index=False)
                
                # VFF Acreedoras sheet
                if not self.vff_acreedoras.empty:
                    self.vff_acreedoras.to_excel(writer, sheet_name='Acreedoras', index=False)
                
                # VFF Debtor Notes (errors - negative differences)
                if not self.vff_debtor_notes.empty:
                    self.vff_debtor_notes.to_excel(writer, sheet_name='VFF_Notas_Deudoras', index=False)
                
                # VFF Fatal Errors that caused crashes during matches
                if not self.vff_abnormal.empty:
                    self.vff_abnormal.to_excel(writer, sheet_name='VFF_Error_Fatal', index=False)

                if not self.merged.empty:
                    self.merged.to_excel(writer, sheet_name='Registro_Auditoria', index=False)
                
                # Build Index Sheet
                self._build_index_sheet(writer)
                
                # Apply styling to all sheets
                self._apply_excel_styling(writer)
                
            print(f"EXITO. Reporte guardado en: {output_file}")
        except PermissionError:
            print(f"ERROR: Cierre {output_file} e intente de nuevo.")

    def _build_index_sheet(self, writer):
        """Creates an Index sheet with hyperlinks to all other sheets."""
        workbook = writer.book
        if 'Indice' in workbook.sheetnames:
            return
            
        index_ws = workbook.create_sheet('Indice', 0)
        workbook.active = index_ws
        
        # Title
        title_cell = index_ws.cell(row=1, column=1, value="INDICE DE HOJAS")
        title_cell.font = Font(bold=True, size=14, color='FFFFFF')
        title_cell.fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        row_idx = 3
        for sheet_name in workbook.sheetnames:
            if sheet_name == 'Indice':
                continue
            
            # Create hyperlink to the sheet
            cell = index_ws.cell(row=row_idx, column=1, value=f"👉 Ir a: {sheet_name}")
            cell.hyperlink = f"#'{sheet_name}'!A1"
            cell.font = Font(color="0563C1", underline="single", bold=True)
            row_idx += 2
            
        index_ws.column_dimensions['A'].width = 40

    def _apply_excel_styling(self, writer):
        """Applies professional formatting to all Excel sheets."""
        # Style definitions
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        subtotal_font = Font(bold=True, size=10)
        subtotal_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        
        alt_fill_1 = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')  # Light blue
        alt_fill_2 = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')   # White
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        workbook = writer.book
        for sheet_name in workbook.sheetnames:
            if sheet_name == 'Indice':
                continue
                
            ws = workbook[sheet_name]
            if ws.max_row is None or ws.max_row < 1:
                continue
            
            # 1. Style headers (row 1)
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # 2. Style data rows
            for row in range(2, ws.max_row + 1):
                first_cell_value = str(ws.cell(row=row, column=1).value or '')
                is_subtotal = first_cell_value.startswith('SUBTOTAL')
                is_inline_header = first_cell_value in ['Nota Deudora', 'Nota Acreedora']
                is_blank = first_cell_value.strip() == '' and all(str(ws.cell(row=row, column=c).value or '').strip() == '' for c in range(1, ws.max_column + 1))
                
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    
                    if is_blank:
                        cell.border = Border() # No border
                        continue # No fill either
                        
                    cell.border = thin_border
                    
                    if is_subtotal:
                        cell.font = subtotal_font
                        cell.fill = subtotal_fill
                    elif is_inline_header:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                    else:
                        # Alternating row colors
                        cell.fill = alt_fill_1 if row % 2 == 0 else alt_fill_2
                        
                    # Format float/int values to 2 decimal accounting format
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00'
            
            # 3. Auto-column-width
            for col in range(1, ws.max_column + 1):
                max_length = 0
                for row in range(1, ws.max_row + 1):
                    cell_value = str(ws.cell(row=row, column=col).value or '')
                    max_length = max(max_length, len(cell_value))
                adjusted_width = min(max_length + 3, 50)  # Cap at 50
                ws.column_dimensions[get_column_letter(col)].width = adjusted_width
            
            # 4. Add "Volver al Indice" link
            back_col = ws.max_column + 2
            back_cell = ws.cell(row=1, column=back_col, value="⬅ Volver al Indice")
            back_cell.hyperlink = "#'Indice'!A1"
            back_cell.font = Font(color="0563C1", underline="single", bold=True)
            ws.column_dimensions[get_column_letter(back_col)].width = 25
            

    def run(self):
        print(f"--- Iniciando Conciliacion en {self.folder_path} ---")
        if not self.load_data(): return
        if not self.validate(): return
        if not self.match_transactions(): return
        self.run_analysis()
        self.generate_reports()
        self.export_results('CONCILIATION_FINAL_REPORT.xlsx')


# =============================================================================
# BACKWARD COMPATIBILITY
# =============================================================================

# This function matches the original function name and signature (if any)
# allowing existing external scripts/tests to call it without breaking.

def robust_conciliation_duplicates_allowed():
    # Also expose internal helper functions for testing if needed, 
    # but primarily this function runs the full flow.
    # Note: Global FOLDER_PATH constant logic is now encapsulated in the Class default
    conciliator = Conciliator(folder_path=DEFAULT_FOLDER_PATH) # Uses the global default constant
    conciliator.run()

# Expose older functions by mapping them to class methods IF necessary for tests.
# Since current tests import specific functions like `check_orphans`, we might 
# need to create wrappers or update tests.
#
# STRATEGY: 
# The implementation below provides functional wrappers that re-use the logic 
# by instantiating a dummy class or just implementing them as standalone 
# helper functions again if we want to support partial testing.
#
# HOWEVER, for cleaner code, we should prefer updating the tests. 
# But to satisfy the "refactor without breaking" constraint as much as possible:

def check_orphans(df_debt, df_credit, merged):
    # Wrapper to support existing unit tests that call this function directly
    c = Conciliator()
    c.df_debt = df_debt
    c.df_credit = df_credit
    c.merged = merged
    return c._check_orphans()

# We can do similar wrappers for others if strictly needed by tests:
# check_intra_pile_duplicates
# validate_files_are_different
# check_data_quality

def check_intra_pile_duplicates(individual_files, label):
    return Conciliator()._check_intra_pile_duplicates(individual_files, label)

def validate_files_are_different(df1, df2):
    return Conciliator()._validate_files_are_different(df1, df2)

def check_data_quality(df, label):
    return Conciliator()._check_data_quality(df, label)


if __name__ == "__main__":
    robust_conciliation_duplicates_allowed()